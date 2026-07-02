import pandas as pd
from openpyxl import load_workbook
import numpy as np

def transfer_data_multiple_sheets(source_file, target_file, sheet_column_map, filters, header_rows, default_columns):
    """
    Transfers specified columns from a source Excel file to specified sheets and columns in a target Excel file,
    with support for unique filters per sheet. Allows population of default columns with conditions.

    Parameters:
    - source_file: Path to the source Excel file.
    - target_file: Path to the target Excel file.
    - sheet_column_map: Dictionary where each key is a target sheet name and the value is a list of tuples 
                        mapping source columns to target columns.
    - filters: A dictionary of filter conditions for each sheet.
    - header_rows: A dictionary where the key is a sheet name and the value is the header row number.
    - default_columns: A dictionary where the key is a sheet name and the value is a list of tuples 
                        mapping target columns to default values and any conditions for applying them.
    """
    
    # Load the source data
    source_data = pd.read_excel(source_file, dtype=str)
    
    # Load the target workbook
    target_wb = load_workbook(target_file)

    for sheet_name, column_map_list in sheet_column_map.items():
        # Ensure the target sheet exists
        if sheet_name not in target_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in target workbook.")
        
        # Select the target sheet
        target_ws = target_wb[sheet_name]
        
        # Get header row and target columns' letters
        header_row = header_rows[sheet_name]
        target_column_letters = get_target_column_letters(target_ws, column_map_list, header_row)

        # Apply filters specific to this sheet
        if filters and sheet_name in filters:
            sheet_data = apply_filters(source_data.copy(), filters[sheet_name])
        else:
            sheet_data = source_data.copy()  # No filters, use all data

        # Define the start row for data writing
        start_row = header_row + 1

        # Write filtered data to target sheet
        write_data_to_target_sheet(sheet_data, target_ws, column_map_list, target_column_letters, start_row)

        # Populate default values in the target sheet with conditions
        populate_default_values(target_ws, default_columns.get(sheet_name, []), header_row)

    # Save the updated target file
    target_wb.save(target_file)
    print(f"Data transferred successfully to {target_file} for all specified sheets!")

def apply_filters(source_data, sheet_filters):
    """
    Applies the given filters to the source data for a specific sheet.

    Parameters:
    - source_data: The source data to be filtered.
    - sheet_filters: A list of filter conditions (tuples) or a single filter condition (tuple).

    Returns:
    - The filtered source data.
    """
    sheet_data = source_data.copy()

    # Check if the filters are a single tuple or a list of tuples
    if isinstance(sheet_filters, list):
        # Apply multiple filter conditions
        for filter_col, filter_val in sheet_filters:
            if filter_col not in sheet_data.columns:
                continue

            if filter_val == "Exclude Blanks":
                # Exclude rows where the column value is NaN or blank
                sheet_data = sheet_data[sheet_data[filter_col].notna()]  # Exclude NaN values
                sheet_data = sheet_data[sheet_data[filter_col].apply(lambda x: str(x).strip() != "")]
            else:
                # Apply the filter if it's a specific value
                sheet_data = sheet_data[sheet_data[filter_col] == filter_val]

    elif isinstance(sheet_filters, tuple):
        # Apply a single filter condition
        filter_col, filter_val = sheet_filters

        if filter_col not in sheet_data.columns:
            return sheet_data  # Return unfiltered data if column is missing

        if filter_val == "Exclude Blanks":
            # Exclude rows where the column value is NaN or blank
            sheet_data = sheet_data[sheet_data[filter_col].notna()]
            sheet_data = sheet_data[sheet_data[filter_col].apply(lambda x: str(x).strip() != "")]
        else:
            # Apply the filter if it's a specific value
            sheet_data = sheet_data[sheet_data[filter_col] == filter_val]

    return sheet_data

def get_target_column_letters(target_ws, column_map_list, header_row):
    """
    Maps target columns to their respective Excel column letters.

    Parameters:
    - target_ws: The target worksheet.
    - column_map_list: List of tuples mapping source columns to target columns.
    - header_row: The row where the headers are located.

    Returns:
    - A dictionary mapping target column names to their respective Excel column letters.
    """
    target_column_letters = {}

    for col in target_ws.iter_cols(min_row=header_row, max_row=header_row):
        header = col[0].value

        for mapping in column_map_list:
            # Support both two-value (original) and three-value (concat/stack) tuples
            source_col, target_col = (
                mapping[:2] if isinstance(mapping, tuple) and len(mapping) >= 2 else mapping
            )

            if header == target_col:
                target_column_letters[target_col] = col[0].column_letter

    # Ensure all target columns were found
    missing_columns = [target_col for _, target_col in (mapping[:2] for mapping in column_map_list)]
    missing_columns = [col for col in missing_columns if col not in target_column_letters]

    if missing_columns:
        raise ValueError(f"Target columns {missing_columns} not found in target sheet header row {header_row}.")

    return target_column_letters


def write_data_to_target_sheet(source_data, target_ws, column_map_list, target_column_letters, start_row):
    """
    Writes data from the source DataFrame to the target sheet.
    - Supports concatenation (multiple source columns into one target column).
    - Supports stacking (multiple source columns into separate rows).
    - Removes duplicate rows at the end.

    Parameters:
    - source_data: The source data to be written.
    - target_ws: The target worksheet.
    - column_map_list: List of tuples mapping source columns (or multiple stacked columns) to target columns.
    - target_column_letters: Dictionary of target column names to Excel column letters.
    - start_row: The row to start writing data from.
    """
    
    
    DATE_COLUMNS = {
        "Effective_Date",
        "Hire_Date",
        "Additional_Jobs_Start_Date",
        "W4_Effective_Date",
        "Most_Recent_Enrollment_Date",
        "Original_Coverage_Begin_Date"
        "Availability_Date",
        "Earliest_Hire_Date",
        "Compensation_Effective_Date",
        "Effective as of",
        "Effective Date"
    }
    
 
    def write_as_text(ws, cell_ref, value, target_col=None):
        if pd.isna(value) or value in ["", " "]:
            return

        if target_col in DATE_COLUMNS:
            parsed_date = pd.to_datetime(value, errors="coerce")
            if pd.notnull(parsed_date):
                ws[cell_ref] = parsed_date.strftime("%Y-%m-%d")
                return

        ws[cell_ref].number_format = "@"
        ws[cell_ref] = str(value)


    current_row = start_row  # Track where to write

    for row_index, row_data in enumerate(source_data.itertuples(index=False), start=start_row):
        max_stack_length = 1  # Default to 1 unless stacking is detected

        # Determine if stacking is needed
        for mapping in column_map_list:
            source_col, target_col, mapping_type = (
                mapping if isinstance(mapping, tuple) and len(mapping) == 3 else (*mapping, "default")
            )

            if isinstance(source_col, tuple) and mapping_type == "stack":
                valid_values = [getattr(row_data, col) for col in source_col if getattr(row_data, col) not in [None, "", " "]]
                max_stack_length = max(max_stack_length, len(valid_values))  # Determine max stack depth

        # Write data (concatenation, stacking, or standard mapping)
        for stack_index in range(max_stack_length):
            row_written = False  # Track if valid data was written

            for mapping in column_map_list:
                source_col, target_col, mapping_type = (
                    mapping if isinstance(mapping, tuple) and len(mapping) == 3 else (*mapping, "default")
                )

                if isinstance(source_col, tuple):
                    if mapping_type == "concat":
                        # Concatenate all values into one string
                        concat_value = " - ".join(
                            str(getattr(row_data, col)) if getattr(row_data, col) not in [None, "", " "] else ""
                            for col in source_col
                        ).strip(" - ")

                        if concat_value:
                            write_as_text(
                                target_ws,
                                f"{target_column_letters[target_col]}{current_row}",
                                concat_value,
                                target_col
                            )
                            row_written = True

                    elif mapping_type == "stack" and stack_index < len(source_col):
                        # Stacking: write each value separately in a new row
                        col = source_col[stack_index]
                        value = getattr(row_data, col)

                        if value not in [None, "", " "]:
                            write_as_text(
                                target_ws,
                                f"{target_column_letters[target_col]}{current_row}",
                                value,
                                target_col
                            )
                            row_written = True

                else:  # Default one-to-one mapping
                    if source_col in source_data.columns and target_col in target_column_letters:
                        value = getattr(row_data, source_col)
                        if value not in [None, "", " "]:
                            write_as_text(
                                target_ws,
                                f"{target_column_letters[target_col]}{current_row}",
                                value,
                                target_col
                            )
                            row_written = True

            if row_written:
                current_row += 1  # Move to next row only if valid data was written

    # ? **Remove Duplicate Rows**
    unique_rows = set()
    duplicate_rows = []

    for row in range(start_row, target_ws.max_row + 1):
        row_data = tuple(
            str(target_ws[f"{target_column_letters[target_col]}{row}"].value).strip().replace('-', '').lower()
            if target_ws[f"{target_column_letters[target_col]}{row}"].value is not None else None
            for _, target_col, *_ in column_map_list  # ? Fix: Properly unpack first two elements only
        )

        if row_data in unique_rows:
            duplicate_rows.append(row)  # Mark duplicate row for deletion
        else:
            unique_rows.add(row_data)

    # Remove duplicate rows in reverse order to avoid shifting issues
    for row in reversed(duplicate_rows):
        target_ws.delete_rows(row)

    print(f"Data successfully written to {target_ws.title}, starting from row {start_row}. Removed {len(duplicate_rows)} duplicate rows.")




def populate_default_values(target_ws, column_defaults, header_row):
    """
    Populates default values in specified columns for each sheet based on conditions.

    Parameters:
    - target_ws: The target worksheet.
    - column_defaults: List of tuples mapping target columns to default values and any conditional columns.
    - header_row: The row where the headers are located.
    """
    start_row = header_row + 1
    max_row = target_ws.max_row

    # Only apply defaults if data exists past the header row
    if max_row <= header_row:
        print(f"No data in '{target_ws.title}' sheet; skipping default population.")
        return


    for target_col, default_value, condition_col in column_defaults:
        # Initialize column letters for target and condition columns
        column_letter = None
        condition_letter = None  # Explicitly reset condition_letter at the start of each loop

        # Limit header search strictly to the specified header row
        for cell in target_ws[header_row]:
            if cell.value == target_col:
                column_letter = cell.column_letter
            if condition_col and cell.value == condition_col:
                condition_letter = cell.column_letter

        # Ensure target column was found
        if column_letter is None:
            raise ValueError(f"Column '{target_col}' not found in sheet.")

        # Log a warning if condition column is not found and skip this default
        if condition_col and condition_letter is None:
            print(f"Warning: Condition column '{condition_col}' not found in header row. Skipping defaults for '{target_col}'.")
            continue

        # Apply default values based on the condition column having data
        for row_index in range(start_row, max_row + 1):
            # Check if condition column has data only if condition_letter is set
            if condition_letter:
                condition_value = target_ws[f"{condition_letter}{row_index}"].value
                # Apply default only if condition column has data
                if condition_value and not pd.isna(condition_value) and condition_value != "":
                    target_ws[f"{column_letter}{row_index}"] = default_value
            elif not condition_col:
                # Apply default value unconditionally if no condition column specified
                target_ws[f"{column_letter}{row_index}"] = default_value



